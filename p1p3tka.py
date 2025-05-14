import requests
import csv
from openpyxl import load_workbook
import io

from kivy.app import App
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.label import Label
from kivy.uix.image import Image
from kivy.uix.scrollview import ScrollView
from kivy.uix.boxlayout import BoxLayout

REPO_URL = "https://github.com/armen78782/DB/"
REPO_OWNER = REPO_URL.split('/')[3]
REPO_NAME = REPO_URL.split('/')[4]

FOLDERS = {
    'OSINT - ПОИСК': 'probiv',
    'БАЗА SBERBANK': 'sberbank',
}

class OSINTApp(App):
    def build(self):
        self.layout = FloatLayout()
        self.folder_choice = list(FOLDERS.keys())[0]

        # Фон
        self.bg = Image(source='elliot.jpg', allow_stretch=True, keep_ratio=False)
        self.layout.add_widget(self.bg)

        # Выпадающий список
        self.spinner = Spinner(
            text=self.folder_choice,
            values=list(FOLDERS.keys()),
            size_hint=(0.6, 0.1),
            pos_hint={'center_x': 0.5, 'top': 0.95}
        )
        self.spinner.bind(text=self.set_folder)
        self.layout.add_widget(self.spinner)

        # Поле ввода
        self.input = TextInput(
            hint_text='Введите слово для поиска и нажмите Enter',
            size_hint=(0.8, 0.1),
            pos_hint={'center_x': 0.5, 'top': 0.8},
            multiline=False
        )
        self.input.bind(on_text_validate=self.on_enter)
        self.layout.add_widget(self.input)

        # Вывод результата (прокручиваемый)
        self.result_label = Label(
            text='',
            markup=True,
            valign='top',
            size_hint_y=None
        )
        self.result_label.bind(texture_size=self.update_label_height)

        self.scroll = ScrollView(
            size_hint=(0.9, 0.5),
            pos_hint={'center_x': 0.5, 'y': 0.05}
        )
        box = BoxLayout(orientation='vertical', size_hint_y=None)
        box.add_widget(self.result_label)
        self.scroll.add_widget(box)
        self.layout.add_widget(self.scroll)

        return self.layout

    def set_folder(self, spinner, text):
        self.folder_choice = text

    def on_enter(self, instance):
        keyword = self.input.text.strip()
        folder_path = FOLDERS[self.folder_choice]
        self.result_label.text = "[color=00ffff]Поиск...[/color]"
        self.input.text = ''
        self.result_label.text = self.github_search(folder_path, keyword)

    def update_label_height(self, instance, value):
        self.result_label.height = self.result_label.texture_size[1]
        self.result_label.text_size = (self.scroll.width * 0.95, None)

    def github_search(self, folder, keyword):
        output = f"[b]Ищем в папке: {folder}[/b]\n"
        try:
            api_url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{folder}"
            response = requests.get(api_url)

            if response.status_code != 200:
                return f"[color=ff0000]Ошибка доступа: {response.status_code}[/color]"

            hits = 0
            for item in response.json():
                file_url = item['download_url']
                name = item['name']
                if name.endswith('.txt'):
                    content = requests.get(file_url).text
                    hits += self.search_txt(content, keyword, output)
                elif name.endswith('.csv'):
                    content = requests.get(file_url).text
                    hits += self.search_csv(content, keyword, output)
                elif name.endswith('.xlsx'):
                    content = requests.get(file_url).content
                    hits += self.search_xlsx(content, keyword, output)

            if hits == 0:
                return f"[color=ff0000]Совпадений не найдено.[/color]"
            return self.result_label.text
        except Exception as e:
            return f"[color=ff0000]Ошибка: {str(e)}[/color]"

    def search_txt(self, content, keyword, output):
        hits = 0
        result = ""
        for i, line in enumerate(content.splitlines(), 1):
            if keyword.lower() in line.lower():
                result += f"[color=00ff00][{i}][/color] {line}\n"
                hits += 1
        self.result_label.text += result
        return hits

    def search_csv(self, content, keyword, output):
        hits = 0
        result = ""
        csv_reader = csv.reader(io.StringIO(content))
        for i, row in enumerate(csv_reader, 1):
            for cell in row:
                if keyword.lower() in cell.lower():
                    result += f"[color=00ff00][{i}][/color] {cell}\n"
                    hits += 1
        self.result_label.text += result
        return hits

    def search_xlsx(self, content, keyword, output):
        hits = 0
        result = ""
        workbook = load_workbook(filename=io.BytesIO(content))
        for sheet in workbook:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and keyword.lower() in str(cell).lower():
                        result += f"[color=00ff00][{sheet.title}][/color] {str(cell)}\n"
                        hits += 1
        self.result_label.text += result
        return hits

if __name__ == '__main__':
    OSINTApp().run()
